from __future__ import annotations

import subprocess
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


MYSQL_EXE = Path(r"D:\MySQL\MySQL Server 8.0\bin\mysql.exe")
DB_NAME = "tcm_vsl"
DB_USER = "root"
DB_PASSWORD = "123456"
WORKBOOK_NAME = "consult-report-import.xlsx"
IMPORT_OPERATOR = "mock-import"

REPORT_HEADERS = [
    "report_key",
    "patient_name",
    "gender",
    "age",
    "visit_type",
    "basic_disease",
    "allergy_history_type",
    "allergy_history_detail",
    "past_history_type",
    "past_history_detail",
    "current_medication_type",
    "current_medication_detail",
    "chief_complaint",
    "duration",
    "severity_level",
    "sudden_worse",
    "life_impact",
    "highest_risk_level",
    "status",
    "doctor_advice",
    "doctor_summary",
    "suggested_department",
    "need_offline_visit",
    "need_supplement",
    "create_time",
    "update_time",
    "doctor_name",
]

DISEASE_HEADERS = [
    "report_key",
    "sort_order",
    "disease_code",
    "disease_name",
    "disease_path",
    "risk_level",
    "risk_reason",
]

ATTACHMENT_HEADERS = [
    "report_key",
    "file_type",
    "file_name",
    "file_url",
    "file_size",
]


def sql_literal(value):
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, (int, float)):
        return str(int(value) if isinstance(value, bool) or float(value).is_integer() else value)
    if isinstance(value, datetime):
        return "'" + value.strftime("%Y-%m-%d %H:%M:%S") + "'"
    text = str(value).strip()
    if text == "":
        return "NULL"
    text = text.replace("\\", "\\\\").replace("'", "''").replace("\r\n", "\n").replace("\r", "\n")
    return "'" + text + "'"


def normalize_bool(value):
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "是"}


def normalize_text(value, fallback=None):
    if value is None:
        return fallback
    text = str(value).strip()
    return text if text else fallback


def normalize_datetime(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return None
    for pattern in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    raise ValueError(f"无法识别时间格式: {text}")


def locate_workbook():
    matches = list(Path("D:/").rglob(WORKBOOK_NAME))
    if not matches:
        raise FileNotFoundError(f"未找到 {WORKBOOK_NAME}")
    return matches[0]


def load_sheet_rows(ws, headers):
    actual_headers = [ws.cell(row=1, column=i).value for i in range(1, len(headers) + 1)]
    if actual_headers != headers:
        raise ValueError(f"{ws.title} 表头不匹配。\n期望: {headers}\n实际: {actual_headers}")

    rows = []
    for row_idx in range(2, ws.max_row + 1):
        values = [ws.cell(row=row_idx, column=i).value for i in range(1, len(headers) + 1)]
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        rows.append(dict(zip(headers, values)))
    return rows


def validate_reports(reports):
    seen_keys = set()
    for item in reports:
        report_key = normalize_text(item["report_key"])
        if not report_key:
            raise ValueError("reports 中存在空的 report_key")
        if report_key in seen_keys:
            raise ValueError(f"reports 中 report_key 重复: {report_key}")
        seen_keys.add(report_key)

        for field in ("patient_name", "gender", "age", "visit_type", "chief_complaint", "highest_risk_level", "status", "create_time"):
            if item.get(field) in (None, ""):
                raise ValueError(f"reports/{report_key} 缺少必填字段 {field}")


def build_report_no(report_key):
    report_no = f"MOCK{report_key}"
    if len(report_no) > 32:
        raise ValueError(f"report_key 过长，生成的 report_no 超过 32 位: {report_key}")
    return report_no


def mysql_query(sql):
    command = [
        str(MYSQL_EXE),
        f"-u{DB_USER}",
        f"-p{DB_PASSWORD}",
        "-D",
        DB_NAME,
        "--default-character-set=utf8mb4",
        "-N",
        "-B",
        "-e",
        sql,
    ]
    result = subprocess.run(command, capture_output=True, text=True, check=True)
    return result.stdout


def mysql_execute(sql):
    command = [
        str(MYSQL_EXE),
        f"-u{DB_USER}",
        f"-p{DB_PASSWORD}",
        "-D",
        DB_NAME,
        "--default-character-set=utf8mb4",
    ]
    subprocess.run(command, input=sql.encode("utf-8"), check=True)


def build_import_sql(reports, diseases_by_key, attachments_by_key):
    statements = ["START TRANSACTION;"]

    for item in reports:
        report_key = normalize_text(item["report_key"])
        report_no = build_report_no(report_key)
        create_time = normalize_datetime(item["create_time"])
        update_time = normalize_datetime(item["update_time"]) or create_time
        doctor_name = normalize_text(item["doctor_name"])

        report_values = [
            sql_literal(report_no),
            "NULL",
            sql_literal(normalize_text(item["patient_name"])),
            sql_literal(normalize_text(item["gender"])),
            sql_literal(int(item["age"])),
            sql_literal(normalize_text(item["visit_type"])),
            sql_literal(normalize_text(item["basic_disease"])),
            sql_literal(normalize_text(item["allergy_history_type"], "无")),
            sql_literal(normalize_text(item["allergy_history_detail"])),
            sql_literal(normalize_text(item["past_history_type"], "无")),
            sql_literal(normalize_text(item["past_history_detail"])),
            sql_literal(normalize_text(item["current_medication_type"], "无")),
            sql_literal(normalize_text(item["current_medication_detail"])),
            sql_literal(normalize_text(item["chief_complaint"])),
            sql_literal(normalize_text(item["duration"])),
            sql_literal(int(item["severity_level"]) if item["severity_level"] is not None else None),
            sql_literal(normalize_text(item["sudden_worse"])),
            sql_literal(normalize_text(item["life_impact"])),
            sql_literal(normalize_text(item["highest_risk_level"], "normal")),
            sql_literal(normalize_text(item["status"], "pending")),
            "NULL",
            sql_literal(doctor_name),
            sql_literal(normalize_text(item["doctor_advice"])),
            sql_literal(normalize_text(item["doctor_summary"])),
            sql_literal(normalize_text(item["suggested_department"])),
            sql_literal(normalize_bool(item["need_offline_visit"])),
            sql_literal(normalize_bool(item["need_supplement"])),
            sql_literal(f"mock-import:{report_key}"),
            sql_literal(IMPORT_OPERATOR),
            sql_literal(create_time),
            sql_literal(IMPORT_OPERATOR),
            sql_literal(update_time),
            sql_literal("0"),
        ]

        statements.append(
            "INSERT INTO tcm_consult_report ("
            "report_no, user_id, patient_name, gender, age, visit_type, basic_disease, "
            "allergy_history_type, allergy_history_detail, past_history_type, past_history_detail, "
            "current_medication_type, current_medication_detail, chief_complaint, duration, severity_level, "
            "sudden_worse, life_impact, highest_risk_level, status, doctor_id, doctor_name, doctor_advice, "
            "doctor_summary, suggested_department, need_offline_visit, need_supplement, remark, create_by, "
            "create_time, update_by, update_time, del_flag"
            ") VALUES ("
            + ", ".join(report_values)
            + ");"
        )
        statements.append(f"SET @rid_{report_key} = LAST_INSERT_ID();")

        for disease in diseases_by_key[report_key]:
            disease_values = [
                f"@rid_{report_key}",
                sql_literal(normalize_text(disease["disease_code"])),
                sql_literal(normalize_text(disease["disease_name"])),
                sql_literal(normalize_text(disease["disease_path"])),
                sql_literal(normalize_text(disease["risk_level"], "normal")),
                sql_literal(normalize_text(disease["risk_reason"])),
                sql_literal(int(disease["sort_order"]) if disease["sort_order"] is not None else 0),
                sql_literal(create_time),
            ]
            statements.append(
                "INSERT INTO tcm_consult_report_disease ("
                "report_id, disease_code, disease_name, disease_path, risk_level, risk_reason, sort_order, create_time"
                ") VALUES ("
                + ", ".join(disease_values)
                + ");"
            )

        for attachment in attachments_by_key[report_key]:
            attachment_values = [
                f"@rid_{report_key}",
                sql_literal(normalize_text(attachment["file_type"], "other")),
                sql_literal(normalize_text(attachment["file_name"])),
                sql_literal(normalize_text(attachment["file_url"])),
                sql_literal(int(attachment["file_size"]) if attachment["file_size"] is not None else None),
                "NULL",
                sql_literal(create_time),
            ]
            statements.append(
                "INSERT INTO tcm_consult_report_attachment ("
                "report_id, file_type, file_name, file_url, file_size, remark, create_time"
                ") VALUES ("
                + ", ".join(attachment_values)
                + ");"
            )

    statements.append("COMMIT;")
    return "\n".join(statements)


def ensure_not_imported(report_nos):
    joined = ", ".join(sql_literal(item) for item in report_nos)
    sql = f"SELECT report_no FROM tcm_consult_report WHERE report_no IN ({joined});"
    existing = [line.strip() for line in mysql_query(sql).splitlines() if line.strip()]
    if existing:
        raise RuntimeError("以下 report_no 已存在，已停止导入: " + ", ".join(existing))


def main():
    workbook_path = locate_workbook()
    wb = load_workbook(workbook_path, data_only=True)

    reports = load_sheet_rows(wb["reports"], REPORT_HEADERS)
    diseases = load_sheet_rows(wb["diseases"], DISEASE_HEADERS)
    attachments = load_sheet_rows(wb["attachments"], ATTACHMENT_HEADERS)

    validate_reports(reports)

    report_keys = {normalize_text(item["report_key"]) for item in reports}
    diseases_by_key = defaultdict(list)
    attachments_by_key = defaultdict(list)

    for item in diseases:
        report_key = normalize_text(item["report_key"])
        if report_key not in report_keys:
            raise ValueError(f"diseases 中存在未知 report_key: {report_key}")
        diseases_by_key[report_key].append(item)

    for item in attachments:
        report_key = normalize_text(item["report_key"])
        if report_key not in report_keys:
            raise ValueError(f"attachments 中存在未知 report_key: {report_key}")
        attachments_by_key[report_key].append(item)

    report_nos = [build_report_no(normalize_text(item["report_key"])) for item in reports]
    ensure_not_imported(report_nos)

    sql = build_import_sql(reports, diseases_by_key, attachments_by_key)
    mysql_execute(sql)

    imported_count = mysql_query(
        "SELECT COUNT(*) FROM tcm_consult_report WHERE remark LIKE 'mock-import:%';"
    ).strip()
    print(f"导入完成: reports={len(reports)}, diseases={len(diseases)}, attachments={len(attachments)}")
    print(f"当前库中 mock-import 报告总数: {imported_count}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:  # noqa: BLE001
        print(f"导入失败: {exc}", file=sys.stderr)
        sys.exit(1)
